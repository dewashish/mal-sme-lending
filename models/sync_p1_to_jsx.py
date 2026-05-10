"""
Sync canonical numbers from Mal_P1_SmartInvoice.xlsx → mal/p1-model-data.js

Run after editing the Excel:
  .venv-models/bin/python models/sync_p1_to_jsx.py

The JSX prototype reads window.MAL_P1_DATA — never recomputes the model in JS.
This keeps Excel as single source of truth.
"""
import json, re
import openpyxl
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'models' / 'Mal_P1_SmartInvoice.xlsx'
OUT  = ROOT / 'mal' / 'p1-model-data.js'

wb = openpyxl.load_workbook(XLSX, data_only=True)

def cell(sheet, row, col):
    return wb[sheet].cell(row=row, column=col).value

def num(v, default=0):
    if v is None: return default
    if isinstance(v, (int, float)): return v
    try: return float(str(v).replace(',', ''))
    except: return default

def col_range(sheet, row, cols):
    return [cell(sheet, row, c) for c in cols]

# ---------- Inputs (the editable assumptions) ----------
inp = {}
ws = wb['Inputs']
for r in range(5, 41):
    label = ws.cell(row=r, column=1).value
    val   = ws.cell(row=r, column=2).value
    name  = ws.cell(row=r, column=5).value  # named range key
    if isinstance(name, str) and name.startswith('('): continue
    if isinstance(name, str) and name and val is not None:
        inp[name] = val

# ---------- Risk Layer (PD/LGD/EAD) ----------
risk = {}
for col_name, col_idx in [('pay30', 2), ('bnpl', 3), ('ext', 4)]:
    risk[col_name] = {
        'pd':       num(cell('Risk', 5, col_idx)),
        'lgd':      num(cell('Risk', 6, col_idx)),
        'ead':      num(cell('Risk', 7, col_idx)),
        'netLoss':  num(cell('Risk', 9, col_idx)),
        'stressed': num(cell('Risk', 10, col_idx)),
    }
risk['stressMultiplier'] = num(cell('Risk', 5, 5)) / num(cell('Risk', 5, 2))  # 3 / 2 = 1.5

# ---------- Disbursement Plan (5-yr) ----------
disb = {}
for i, y in enumerate(['y1','y2','y3','y4','y5']):
    col = 2 + i
    disb[y] = {
        'face':           num(cell('Disbursement Plan', 5,  col)),
        'yoyGrowth':      num(cell('Disbursement Plan', 6,  col)),
        'advanced':       num(cell('Disbursement Plan', 8,  col)),
        'avgBook':        num(cell('Disbursement Plan', 10, col)),
        'pay30Share':     num(cell('Disbursement Plan', 13, col)),
        'bnplShare':      num(cell('Disbursement Plan', 14, col)),
        'extShare':       num(cell('Disbursement Plan', 15, col)),
    }

# ---------- Combined P&L (5-yr) ----------
pnl = {}
ratios = {}
for i, y in enumerate(['y1','y2','y3','y4','y5']):
    col = 2 + i
    pnl[y] = {
        'avgBook':       num(cell('Combined P&L', 5,  col)),
        'grossInt':      num(cell('Combined P&L', 6,  col)),
        'fundingCost':   num(cell('Combined P&L', 7,  col)),
        'nim':           num(cell('Combined P&L', 8,  col)),
        'lossProvision': num(cell('Combined P&L', 9,  col)),
        'ram':           num(cell('Combined P&L', 10, col)),
        'opex':          num(cell('Combined P&L', 11, col)),
        'netContrib':    num(cell('Combined P&L', 12, col)),
        'equity':        num(cell('Combined P&L', 13, col)),
    }
    ratios[y] = {
        'nimPct':        num(cell('Combined P&L', 16, col)),
        'lossPct':       num(cell('Combined P&L', 17, col)),
        'opexPct':       num(cell('Combined P&L', 18, col)),
        'netMarginPct':  num(cell('Combined P&L', 19, col)),
        'roaPct':        num(cell('Combined P&L', 20, col)),
        'rarocPct':      num(cell('Combined P&L', 21, col)),
    }

# ---------- Capital Plan (5-yr) ----------
cap = {}
for i, y in enumerate(['y1','y2','y3','y4','y5']):
    col = 2 + i
    cap[y] = {
        'avgBook':       num(cell('Capital Plan', 6,  col)),
        'rwa':           num(cell('Capital Plan', 7,  col)),
        'requiredEq':    num(cell('Capital Plan', 9,  col)),
        'raisedYear':    num(cell('Capital Plan', 11, col)),
        'cumulativeRaised':num(cell('Capital Plan', 12, col)),
        'netContribAnn': num(cell('Capital Plan', 14, col)),
        'retainedEarnings':num(cell('Capital Plan', 15, col)),
        'totalAvailEq':  num(cell('Capital Plan', 17, col)),
        'headroom':      num(cell('Capital Plan', 19, col)),
        'tier1Ratio':    num(cell('Capital Plan', 21, col)),
        'tier1Status':   cell('Capital Plan', 22, col) or '',
    }

# ---------- Funding Stack ----------
fund = {
    'tranches': [],
    'blendedCof':   num(cell('Funding Stack', 10, 2)),
    'inputsCof':    num(cell('Funding Stack', 11, 2)),
    'covenants':    {},
    'waterfall':    {},
}
for r in range(5, 8):
    fund['tranches'].append({
        'name':    cell('Funding Stack', r, 1) or '',
        'share':   num(cell('Funding Stack', r, 2)),
        'spread':  num(cell('Funding Stack', r, 3)),
        'eibor':   num(cell('Funding Stack', r, 4)),
        'allIn':   num(cell('Funding Stack', r, 5)),
    })
for i, y in enumerate(['y1','y2','y3','y4','y5']):
    col = 2 + i
    fund['waterfall'][y] = {
        'totalBook':  num(cell('Funding Stack', 16, col)),
        'senior':     num(cell('Funding Stack', 17, col)),
        'mezz':       num(cell('Funding Stack', 18, col)),
        'equity':     num(cell('Funding Stack', 19, col)),
    }
fund['covenants'] = {
    'seniorCap':       [cell('Funding Stack', 22, c) or '' for c in range(2, 7)],
    'singleObligor':   [cell('Funding Stack', 23, c) or '' for c in range(2, 7)],
    'tier1Floor':      [cell('Funding Stack', 24, c) or '' for c in range(2, 7)],
}

# ---------- Scenarios (4-way Y3 KPI panel) ----------
scen = {}
SCEN_COLS = {'base': 3, 'upside': 4, 'downside': 5, 'stress': 6}
for s, col in SCEN_COLS.items():
    scen[s] = {
        'disbY3':        num(cell('Scenarios', 19, col)),
        'pay30Book':     num(cell('Scenarios', 20, col)),
        'bnplBook':      num(cell('Scenarios', 21, col)),
        'extBook':       num(cell('Scenarios', 22, col)),
        'totalBook':     num(cell('Scenarios', 23, col)),
        'grossInt':      num(cell('Scenarios', 25, col)),
        'fundingCost':   num(cell('Scenarios', 26, col)),
        'nim':           num(cell('Scenarios', 27, col)),
        'loss':          num(cell('Scenarios', 28, col)),
        'opex':          num(cell('Scenarios', 29, col)),
        'netContrib':    num(cell('Scenarios', 30, col)),
        'roaPct':        num(cell('Scenarios', 31, col)),
        'equity':        num(cell('Scenarios', 32, col)),
        'rarocPct':      num(cell('Scenarios', 33, col)),
    }

# ---------- Break-even ----------
be = {
    'y2BE':            num(cell('Break-even', 18, 2)),
    'y2Plan':          num(cell('Break-even', 19, 2)),
    'y2Headroom':      num(cell('Break-even', 20, 2)),
    'y2Status':        cell('Break-even', 21, 2) or '',
    'pdMultiplier':    num(cell('Break-even', 34, 2)),
    'pdShift':         cell('Break-even', 35, 2) or '',
    'vsStress':        cell('Break-even', 36, 2) or '',
    'vsDownside':      cell('Break-even', 37, 2) or '',
    'targetROA':       num(cell('Break-even', 26, 2)),
    'allowableLoss':   num(cell('Break-even', 33, 2)),
}

# ---------- Vintage Cohorts ----------
vintage = {
    'lifetimeLossPct': num(cell('Vintage Cohorts', 11, 4)),
    'totalAdvanced':   num(cell('Vintage Cohorts', 14, 3)),
    'totalLifetimeLoss': num(cell('Vintage Cohorts', 15, 5)),
    'perVintage': {},
    'lossCurve': [],
}
for i, y in enumerate(['y1', 'y2', 'y3']):
    r = 11 + i
    vintage['perVintage'][y] = {
        'disbursed': num(cell('Vintage Cohorts', r, 2)),
        'advanced':  num(cell('Vintage Cohorts', r, 3)),
        'lossPct':   num(cell('Vintage Cohorts', r, 4)),
        'lossM':     num(cell('Vintage Cohorts', r, 5)),
    }
# Loss curve M0-M24 (row 21, cols 2-26)
months = []
cumPct = []
for col in range(2, 27):
    label = cell('Vintage Cohorts', 20, col)
    pct   = cell('Vintage Cohorts', 21, col)
    if label is not None and pct is not None:
        months.append(str(label))
        cumPct.append(num(pct))
vintage['lossCurve'] = {'months': months, 'cumPct': cumPct}

# Per-vintage cumulative AED M curves (rows 26/27/28)
vintage['lossCurveAED'] = {}
for i, y in enumerate(['y1', 'y2', 'y3']):
    r = 26 + i
    vals = []
    for col in range(2, 27):
        v = cell('Vintage Cohorts', r, col)
        if v is not None: vals.append(num(v))
    vintage['lossCurveAED'][y] = vals

# ---------- Sensitivity grid (Dashboard mirror) ----------
sens = {
    'cofLevels': [],
    'pdMultipliers': [],
    'grid': [],
}
# Header row at row 30, col 3+ (PD multipliers)
for col in range(3, 9):
    v = cell('Dashboard', 30, col)
    if v is not None:
        m = re.match(r'([\d.]+)', str(v))
        if m: sens['pdMultipliers'].append(float(m.group(1)))
# COF labels at rows 31-37, col 2
for r in range(31, 38):
    v = cell('Dashboard', r, 2)
    if v is not None:
        m = re.match(r'COF ([\d.]+)', str(v))
        if m: sens['cofLevels'].append(float(m.group(1)))
# Grid values
for r in range(31, 38):
    row = []
    for col in range(3, 9):
        row.append(num(cell('Dashboard', r, col)))
    if any(x for x in row):
        sens['grid'].append(row)

# ---------- Unit Economics (per 100k, by track) ----------
ue = {}
for col_name, col_idx in [('pay30', 2), ('bnpl', 3), ('ext', 4)]:
    ue[col_name] = {
        'face':           num(cell('Unit Economics', 5,  col_idx)),
        'advanced':       num(cell('Unit Economics', 6,  col_idx)),
        'tenor':          num(cell('Unit Economics', 7,  col_idx)),
        'yieldPct':       num(cell('Unit Economics', 8,  col_idx)),
        'grossFee':       num(cell('Unit Economics', 9,  col_idx)),
        'fundingCost':    num(cell('Unit Economics', 10, col_idx)),
        'nim':            num(cell('Unit Economics', 11, col_idx)),
        'expectedLoss':   num(cell('Unit Economics', 12, col_idx)),
        'ram':            num(cell('Unit Economics', 13, col_idx)),
        'opex':           num(cell('Unit Economics', 14, col_idx)),
        'netContrib':     num(cell('Unit Economics', 15, col_idx)),
        'netContribPct':  num(cell('Unit Economics', 16, col_idx)),
    }

# ---------- Track P&L (per track 5-yr) ----------
tracks = {}
for sheet, key in [('Track A · Pay-30', 'pay30'),
                   ('Track B · BNPL', 'bnpl'),
                   ('Track C · Extension', 'ext')]:
    tracks[key] = {}
    for i, y in enumerate(['y1','y2','y3','y4','y5']):
        col = 2 + i
        tracks[key][y] = {
            'avgBook':     num(cell(sheet, 5,  col)),
            'yieldPct':    num(cell(sheet, 6,  col)),
            'grossInt':    num(cell(sheet, 7,  col)),
            'fundingCost': num(cell(sheet, 10, col)),
            'nim':         num(cell(sheet, 12, col)),
            'lossPct':     num(cell(sheet, 16, col)),
            'loss':        num(cell(sheet, 17, col)),
            'ram':         num(cell(sheet, 19, col)),
            'opex':        num(cell(sheet, 22, col)),
            'netContrib':  num(cell(sheet, 24, col)),
            'roaPct':      num(cell(sheet, 25, col)),
            'rarocPct':    num(cell(sheet, 28, col)),
        }

# ---------- Aggregates / headlines ----------
fiveYrSum = {
    'face':            sum(disb[y]['face'] for y in ['y1','y2','y3','y4','y5']),
    'avgBookPeak':     pnl['y5']['avgBook'],
    'cumNetContrib':   sum(pnl[y]['netContrib'] for y in ['y1','y2','y3','y4','y5']),
    'cumGrossInt':     sum(pnl[y]['grossInt'] for y in ['y1','y2','y3','y4','y5']),
    'cumLoss':         sum(pnl[y]['lossProvision'] for y in ['y1','y2','y3','y4','y5']),
    'cumOpex':         sum(pnl[y]['opex'] for y in ['y1','y2','y3','y4','y5']),
    'totalEquityRaised': sum(cap[y]['raisedYear'] for y in ['y1','y2','y3','y4','y5']),
    'lifetimeLoss':    vintage['totalLifetimeLoss'],
}

# ---------- Assemble & write ----------
data = {
    'meta': {
        'workbook': XLSX.name,
        'sheets':   len(wb.sheetnames),
        'lastUpdated': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'currency': 'AED',
        'unit':     'M (millions)',
    },
    'inputs': inp,
    'risk':   risk,
    'disbursement': disb,
    'pnl':    pnl,
    'ratios': ratios,
    'capitalPlan': cap,
    'fundingStack': fund,
    'scenarios': scen,
    'breakEven': be,
    'vintage': vintage,
    'sensitivity': sens,
    'unitEconomics': ue,
    'tracks': tracks,
    'fiveYearSummary': fiveYrSum,
}

js = '/* eslint-disable */\n'
js += '// AUTO-GENERATED from models/Mal_P1_SmartInvoice.xlsx — do not edit by hand.\n'
js += f'// Sync command:  .venv-models/bin/python models/sync_p1_to_jsx.py\n'
js += f'// Last sync: {data["meta"]["lastUpdated"]}\n'
js += 'window.MAL_P1_DATA = ' + json.dumps(data, indent=2) + ';\n'

OUT.write_text(js)
print(f'Wrote {OUT}  ({OUT.stat().st_size // 1024} KB)')
print(f'Top-level keys: {list(data.keys())}')
print(f'5-yr cumulative net contribution: AED {fiveYrSum["cumNetContrib"]:.1f}M')
print(f'Total equity raised:              AED {fiveYrSum["totalEquityRaised"]:.0f}M')
print(f'Y3 ROA: {ratios["y3"]["roaPct"]:.2f}%  ·  Y3 RAROC: {ratios["y3"]["rarocPct"]:.1f}%')
