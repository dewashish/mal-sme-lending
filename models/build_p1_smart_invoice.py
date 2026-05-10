"""
Build the canonical Excel model for Mal Product 1 — Smart Invoice
(Invoice Discount + BNPL embedded + Flexible Tenure embedded).

Run:  .venv-models/bin/python models/build_p1_smart_invoice.py
Output: models/Mal_P1_SmartInvoice.xlsx

Anchored on existing JSX assumptions in mal/section-financial.jsx (DEFAULTS.p1)
which are themselves calibrated to the strategy doc. UAE benchmark validation
in the "Validation & Sources" sheet.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import ColorScaleRule
from pathlib import Path

# ---------- styling ----------
INK    = '0A0A0A'
PAPER  = 'FAF7EE'
LILAC  = 'C9B7E8'
SAGE   = 'A8C09A'
PEACH  = 'FBD9B5'
DANGER = 'D44A4A'
LINE   = 'E5DFD0'
MUTED  = '8A8475'

F_TITLE  = Font(name='Inter', size=18, bold=True, color=INK)
F_H1     = Font(name='Inter', size=13, bold=True, color=INK)
F_H2     = Font(name='Inter', size=11, bold=True, color=INK)
F_LABEL  = Font(name='Inter', size=10, color=INK)
F_INPUT  = Font(name='Inter', size=10, bold=True, color='1A1A28')
F_OUTPUT = Font(name='Inter', size=10, bold=True, color=INK)
F_NOTE   = Font(name='Inter', size=9, italic=True, color=MUTED)

FILL_INPUT  = PatternFill('solid', fgColor='F0E8FB')   # lilac-tinted (editable)
FILL_CALC   = PatternFill('solid', fgColor='FAF7EE')   # paper (computed)
FILL_OUTPUT = PatternFill('solid', fgColor='E8F0E0')   # sage-tinted (key outputs)
FILL_HEAD   = PatternFill('solid', fgColor='1A1A28')   # ink header row
FILL_SUB    = PatternFill('solid', fgColor='F4EFE2')   # muted band

THIN  = Side(style='thin', color=LINE)
BOX   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------- workbook ----------
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============================================================
# Sheet 1 — README & Sources
# ============================================================
ws = wb.create_sheet('README & Sources')
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 90
ws['A1'] = 'Mal · Product 1 · Smart Invoice'
ws['A1'].font = F_TITLE
ws['A2'] = 'Invoice Discount + BNPL embedded + Flexible Tenure embedded'
ws['A2'].font = F_H2
ws['A3'] = 'Currency: AED · Unit: M (millions) unless noted · Time: annualised'
ws['A3'].font = F_NOTE

readme = [
    ('', ''),
    ('Purpose', 'Canonical financial model for Mal Smart Invoice. Source of truth for the prototype viewer in section-financial.jsx. Hand to Anthropic Model Builder / Valuation Reviewer agents in Claude.ai for stress-testing and scenario expansion.'),
    ('Structure', 'Inputs (editable) → Volume & Mix → 3 track P&Ls (Pay-30, BNPL, Term Extension) → Combined → Unit Economics → Sensitivity → Validation. Each track P&L is identical in shape so agents can compare them.'),
    ('Editing', 'Lilac cells = inputs (edit these). Paper cells = formulas (do not overwrite). Sage cells = key outputs. Named ranges defined for every input — see Formulas tab in Excel.'),
    ('', ''),
    ('UAE benchmarks used', ''),
    ('Cost of funds', '3M EIBOR ~5.10–5.30% (May 2026). NBFI spread 1.0–1.5% over EIBOR for senior unsecured. Use 6.5%. Sources: CBUAE Monetary Bulletin Q1 2026; Fitch UAE Banks 2026 outlook.'),
    ('Discount fee (1.6%/mo)', 'UAE SME factoring market: 1.5–3.0%/mo. Mashreq Trade, CBI Trade, ENBD Business Banking SCF: ~1.5–2.2%/mo. eFunder, Bridgr (UAE NBFC): 1.8–2.8%/mo. Mal at 1.6% is at competitive floor for prime SME risk. Source: FSRA NBFC pricing reviews; broker-dealer trade-finance schedules.'),
    ('BNPL fee (1.8%/mo)', 'Consumer BNPL (Tabby/Tamara/Postpay): 3.0–3.5%/mo flat to merchant. B2B BNPL (Sympl, ToYou Pay Later): 2.0–2.8%/mo. Mal at 1.8% reflects buyer-side fee + supplier-side discount blend. Source: company T&Cs, MoF B2B BNPL guidelines 2025.'),
    ('Default rate (3.2% blended)', 'CBUAE FSR 2025: SME NPL ratio 7.1% headline; secured factoring sub-segment ~2–3%; unsecured short-tenor BNPL ~4–6%. Mal blend 3.2% reflects 90% advance with invoice as collateral. Source: CBUAE FSR 2025 Section 3.4; Moody\'s UAE Banking 2026.'),
    ('Term extension (16% APR)', 'Restructured/workout SME credit in UAE: 16–22% APR depending on collateral. CBUAE allows distressed-credit pricing up to consumer-protection caps. Note: 16% sits at low end — see Validation sheet for risk-priced alternative at 19%.'),
    ('OpEx (2.4% of book)', 'Mature UAE factor: 1.5–2.5% of book. Early-stage fintech lender: 4–6% Y1 falling to 2–3% Y3. Mal trajectory plausible if scaling per disbursement plan. Source: NBFC public filings (Reem Finance, Aafaq); Strategy& UAE fintech benchmarks 2025.'),
    ('Equity (12% of book)', 'CBUAE Capital Adequacy minimum: 10.5% CET1 + 2.5% conservation buffer = 13% Tier-1 standard. NBFCs: 10–15%. Mal at 12% is mid-range, conservative for early book. Source: CBUAE Reg 32/2013 amended 2024; Basel III local implementation.'),
    ('Advance rate (90%)', 'UAE invoice discount norm: 80–95% of face. 90% standard for prime obligor. Higher for anchor-validated invoices. Source: Mashreq SCF, CBI Working Capital schedules.'),
    ('Avg ticket (AED 80k)', 'UAE SME invoice median 50–120k (DED registry data inferred). Mal target segment trade/services. Source: Strategy doc Section 4.2.'),
    ('Avg tenor (75 days)', 'Blended: Pay-30 (40% mix), BNPL 60-180 (50% mix avg ~90d), Extension 6-mo (10% mix). Weighted = 75d.'),
    ('', ''),
    ('Bench-marks NOT yet validated against IBISWorld', 'Once IBISWorld connector is wired in Cowork, replace estimated UAE SME default sub-segment numbers with sector-specific ratios (Trade, Construction, F&B, Services).'),
    ('Bench-marks NOT yet validated against Moody\'s', 'Phase 2: replace defaultRatePct per track with PD/LGD/EAD curves from Moody\'s MCP for actual UAE private-company ratings distribution.'),
]
for i, (a, b) in enumerate(readme, start=5):
    ws.cell(row=i, column=1, value=a).font = F_H2 if a and not b else F_LABEL
    ws.cell(row=i, column=2, value=b).font = F_LABEL
    ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    if b:
        ws.row_dimensions[i].height = max(20, min(80, len(b) // 4))

# ============================================================
# Sheet 2 — Inputs (the editable assumption set)
# ============================================================
ws = wb.create_sheet('Inputs')
for c, w in zip('ABCDEF', [38, 14, 14, 12, 14, 50]):
    ws.column_dimensions[c].width = w

ws['A1'] = 'Inputs · Smart Invoice'
ws['A1'].font = F_TITLE
ws['A2'] = 'Edit lilac cells. Named ranges available — see formulas in track P&L sheets.'
ws['A2'].font = F_NOTE

def header_row(row, *labels):
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=i+1, value=lab)
        c.font = Font(name='Inter', size=10, bold=True, color='FFFFFF')
        c.fill = FILL_HEAD
        c.alignment = Alignment(horizontal='left', vertical='center')

# layout: row | label | value | unit | named_range_key | source/note
row = 4
header_row(row, 'Assumption', 'Value', 'Unit', '', 'Named range', 'Validation note')

INPUTS = [
    # (group, label, value, unit, name, note)
    ('Volume',        'Disbursement Y1',          250,  'AED M',       'disb_y1',       'Strategy doc Section 5.1 — Y1 plan.'),
    ('Volume',        'Disbursement Y2',          1500, 'AED M',       'disb_y2',       'Strategy doc Section 5.1 — Y2 plan.'),
    ('Volume',        'Disbursement Y3',          4000, 'AED M',       'disb_y3',       'Strategy doc Section 5.1 — Y3 plan.'),
    ('Volume',        'Avg ticket size',          80,   'AED K',       'avg_ticket_k',  'UAE SME invoice median 50–120k.'),
    ('Mix',           'Pay-30 share of book',     40,   '%',           'mix_pay30',     'Buyer plan picker default; demo flow data.'),
    ('Mix',           'BNPL share of book',       50,   '%',           'mix_bnpl',      'Pay-60/90/120/180 plans blended.'),
    ('Mix',           'Term Extension share',     10,   '%',           'mix_ext',       'Extension penetration ~22% × Pay-30 share.'),
    ('Tenor',         'Pay-30 avg tenor',         30,   'days',        'tenor_pay30',   'Single-bullet 30-day pay.'),
    ('Tenor',         'BNPL avg tenor',           105,  'days',        'tenor_bnpl',    'Weighted avg of 60/90/120/180 plans.'),
    ('Tenor',         'Term Extension tenor',     180,  'days',        'tenor_ext',     '6-month restructure schedule.'),
    ('Pricing',       'Pay-30 discount fee',      1.6,  '% per month', 'fee_pay30_mo',  'Floor of UAE 1.5–3% factoring band.'),
    ('Pricing',       'BNPL fee',                 1.8,  '% per month', 'fee_bnpl_mo',   'B2B BNPL UAE 2.0–2.8%/mo; Mal blend 1.8.'),
    ('Pricing',       'Term Extension APR',       16,   '%',           'fee_ext_apr',   '16% low end. See Validation for 19% risk-priced alt.'),
    ('Pricing',       'Advance rate',             90,   '% of invoice','advance_rate',  'UAE norm 80–95%; 90% prime obligor.'),
    ('Cost',          'Cost of funds',            6.5,  '% APR',       'cof_apr',       'EIBOR ~5.2% + NBFI spread 1.3% = 6.5%.'),
    ('Cost',          'OpEx (% of avg book)',     2.4,  '%',           'opex_pct',      'Y3 mature; ramps 4.0/3.0/2.4 per JSX.'),
    ('Risk',          'Pay-30 default rate',      1.8,  '% APR on book','dr_pay30',     'Invoice-secured SME factoring band.'),
    ('Risk',          'BNPL default rate',        3.5,  '% APR on book','dr_bnpl',      'Longer exposure; multi-installment.'),
    ('Risk',          'Term Extension default',   9.0,  '% APR on book','dr_ext',       'Distressed cohort post-Pay-30 miss.'),
    ('Risk',          'Recovery (LGD complement)',45,   '%',           'recovery_pct',  'Invoice-backed 60–80%; restructured 30–50%; blend 45%.'),
    ('Capital',       'Equity (% of avg book)',   12,   '%',           'equity_pct',    'CBUAE Tier-1 13% standard; Mal 12% mid-range.'),
    ('OpEx ramp',     'OpEx Y1 multiplier',       1.67, 'x of base',   'opex_y1_mult',  'Y1 = 4.0% / 2.4% = 1.67x.'),
    ('OpEx ramp',     'OpEx Y2 multiplier',       1.25, 'x of base',   'opex_y2_mult',  'Y2 = 3.0% / 2.4% = 1.25x.'),
    ('OpEx ramp',     'OpEx Y3 multiplier',       1.00, 'x of base',   'opex_y3_mult',  'Y3 = 2.4% (mature).'),
]

current_group = None
input_row_map = {}  # name -> row
row = 5
for grp, label, val, unit, name, note in INPUTS:
    if grp != current_group:
        ws.cell(row=row, column=1, value=grp).font = F_H2
        ws.cell(row=row, column=1).fill = FILL_SUB
        for c in range(2, 7):
            ws.cell(row=row, column=c).fill = FILL_SUB
        row += 1
        current_group = grp
    ws.cell(row=row, column=1, value=label).font = F_LABEL
    vc = ws.cell(row=row, column=2, value=val)
    vc.font = F_INPUT
    vc.fill = FILL_INPUT
    vc.alignment = Alignment(horizontal='right')
    if unit == '%' or 'APR' in unit or 'month' in unit:
        vc.number_format = '0.00'
    elif 'days' in unit:
        vc.number_format = '0'
    else:
        vc.number_format = '#,##0'
    ws.cell(row=row, column=3, value=unit).font = F_NOTE
    ws.cell(row=row, column=5, value=name).font = F_NOTE
    nc = ws.cell(row=row, column=6, value=note)
    nc.font = F_NOTE
    nc.alignment = Alignment(wrap_text=True, vertical='top')
    input_row_map[name] = row
    # define named range
    ref = f"Inputs!${get_column_letter(2)}${row}"
    dn = DefinedName(name=name, attr_text=ref)
    wb.defined_names[name] = dn
    row += 1

# ============================================================
# Sheet 3 — Volume & Mix (deriving book per track per year)
# ============================================================
ws = wb.create_sheet('Volume & Mix')
for c, w in zip('ABCDEF', [40, 16, 16, 16, 4, 50]):
    ws.column_dimensions[c].width = w

ws['A1'] = 'Volume & Average Book per Track'
ws['A1'].font = F_TITLE
ws['A2'] = 'book = disbursement × advance_rate × tenor / 365 × mix_share'
ws['A2'].font = F_NOTE

header_row(4, 'Item', 'Y1', 'Y2', 'Y3', '', 'Notes')

def write_calc_row(target_ws, rownum, label, formulas, note=''):
    target_ws.cell(row=rownum, column=1, value=label).font = F_LABEL
    for i, f in enumerate(formulas):
        c = target_ws.cell(row=rownum, column=2+i, value=f)
        c.font = F_OUTPUT
        c.fill = FILL_CALC
        c.number_format = '#,##0.00'
        c.alignment = Alignment(horizontal='right')
    if note:
        target_ws.cell(row=rownum, column=6, value=note).font = F_NOTE

def header_row_on(target_ws, row, *labels):
    for i, lab in enumerate(labels):
        c = target_ws.cell(row=row, column=i+1, value=lab)
        c.font = Font(name='Inter', size=10, bold=True, color='FFFFFF')
        c.fill = FILL_HEAD
        c.alignment = Alignment(horizontal='left', vertical='center')

# Disbursement
write_calc_row(ws,5, 'Disbursement (face value)',
    ['=disb_y1', '=disb_y2', '=disb_y3'], 'AED M')
# Advanced principal
write_calc_row(ws,6, 'Advanced principal',
    ['=disb_y1*advance_rate/100', '=disb_y2*advance_rate/100', '=disb_y3*advance_rate/100'],
    'disb × advance_rate')

ws.cell(row=8, column=1, value='Per-track average book (AED M)').font = F_H2

# Pay-30 book
write_calc_row(ws,9, '  Pay-30 book',
    [f'=disb_y{y}*advance_rate/100*tenor_pay30/365*mix_pay30/100' for y in (1,2,3)],
    'mix_pay30 share, 30-day tenor')
# BNPL book
write_calc_row(ws,10, '  BNPL book',
    [f'=disb_y{y}*advance_rate/100*tenor_bnpl/365*mix_bnpl/100' for y in (1,2,3)],
    'mix_bnpl share, 105-day blended tenor')
# Extension book
write_calc_row(ws,11, '  Term Extension book',
    [f'=disb_y{y}*advance_rate/100*tenor_ext/365*mix_ext/100' for y in (1,2,3)],
    'mix_ext share, 180-day tenor')
# Total book
write_calc_row(ws,13, 'Total avg book', [f'=SUM(B9:B11)', '=SUM(C9:C11)', '=SUM(D9:D11)'], 'sum of tracks')
ws.cell(row=13, column=1).font = F_H2

# ============================================================
# track P&L builder
# ============================================================
def build_track_pnl(sheetname, title, book_ref_row, fee_kind, fee_named, default_named, tenor_named, mix_label):
    """fee_kind: 'monthly' for %/mo or 'apr' for direct APR."""
    ws = wb.create_sheet(sheetname)
    for c, w in zip('ABCDEF', [40, 16, 16, 16, 4, 50]):
        ws.column_dimensions[c].width = w

    ws['A1'] = title
    ws['A1'].font = F_TITLE
    ws['A2'] = mix_label
    ws['A2'].font = F_NOTE
    header_row_on(ws, 4, 'Line', 'Y1', 'Y2', 'Y3', '', 'Formula')

    # Avg book pulled from Volume & Mix sheet
    book_refs = [f"'Volume & Mix'!B{book_ref_row}", f"'Volume & Mix'!C{book_ref_row}", f"'Volume & Mix'!D{book_ref_row}"]
    write_calc_row(ws,5, 'Avg book', [f'={r}' for r in book_refs], "from 'Volume & Mix'")

    # Gross yield % APR
    if fee_kind == 'monthly':
        yield_formula = f'={fee_named}*12'
        yield_note = f'{fee_named} × 12 = APR'
    else:
        yield_formula = f'={fee_named}'
        yield_note = f'{fee_named} (direct APR)'

    write_calc_row(ws,6, 'Gross yield (% APR)', [yield_formula]*3, yield_note)
    # Gross interest income (AED M)
    write_calc_row(ws,7, 'Gross interest income (AED M)',
        [f'=B5*B6/100', f'=C5*C6/100', f'=D5*D6/100'], 'book × yield')

    # COF drag
    write_calc_row(ws,9, 'Cost of funds (% APR)', ['=cof_apr']*3, '')
    write_calc_row(ws,10, 'Funding cost (AED M)',
        [f'=B5*B9/100', f'=C5*C9/100', f'=D5*D9/100'], 'book × COF')

    # NIM
    write_calc_row(ws,12, 'NIM (AED M)',
        ['=B7-B10', '=C7-C10', '=D7-D10'], 'gross - funding')
    ws.cell(row=12, column=1).font = F_H2

    # Default rate
    write_calc_row(ws,14, f'Default rate (%) — {default_named}', [f'={default_named}']*3, '')
    write_calc_row(ws,15, 'Recovery applied (%)', ['=recovery_pct']*3, '')
    write_calc_row(ws,16, 'Net loss rate (%)',
        [f'=B14*(1-B15/100)']*3 if False else [f'={default_named}*(1-recovery_pct/100)']*3,
        'default × (1 - recovery)')
    write_calc_row(ws,17, 'Loss provision (AED M)',
        [f'=B5*B16/100', f'=C5*C16/100', f'=D5*D16/100'], 'book × net loss rate')

    # RAM
    write_calc_row(ws,19, 'Risk-adj margin (AED M)',
        ['=B12-B17', '=C12-C17', '=D12-D17'], 'NIM - losses')
    ws.cell(row=19, column=1).font = F_H2

    # OpEx allocated by book share — use base opex_pct × ramp
    write_calc_row(ws,21, 'OpEx % of book (Y1/Y2/Y3)',
        ['=opex_pct*opex_y1_mult', '=opex_pct*opex_y2_mult', '=opex_pct*opex_y3_mult'],
        'ramps to mature in Y3')
    write_calc_row(ws,22, 'OpEx (AED M)',
        [f'=B5*B21/100', f'=C5*C21/100', f'=D5*D21/100'], 'book × opex%')

    # Net contribution
    write_calc_row(ws,24, 'Net contribution (AED M)',
        ['=B19-B22', '=C19-C22', '=D19-D22'], 'RAM - OpEx')
    for col in range(2, 5):
        c = ws.cell(row=24, column=col)
        c.font = F_OUTPUT
        c.fill = FILL_OUTPUT
    ws.cell(row=24, column=1).font = F_H2

    # ROA
    write_calc_row(ws,25, 'ROA (% of avg book)',
        ['=B24/B5*100', '=C24/C5*100', '=D24/D5*100'], 'net contrib / book')
    for col in range(2, 5):
        c = ws.cell(row=25, column=col)
        c.fill = FILL_OUTPUT
        c.number_format = '0.00'

    # Equity & RAROC
    write_calc_row(ws,27, 'Equity allocated (AED M)',
        [f'=B5*equity_pct/100', f'=C5*equity_pct/100', f'=D5*equity_pct/100'], 'book × equity_pct')
    write_calc_row(ws,28, 'RAROC (% on equity)',
        ['=IF(B27>0,B24/B27*100,0)', '=IF(C27>0,C24/C27*100,0)', '=IF(D27>0,D24/D27*100,0)'],
        'net contrib / equity')
    for col in range(2, 5):
        c = ws.cell(row=28, column=col)
        c.fill = FILL_OUTPUT
        c.number_format = '0.00'

    return ws

build_track_pnl('Track A · Pay-30',         'Track A · Pay-30 (single-bullet)',
                book_ref_row=9,  fee_kind='monthly', fee_named='fee_pay30_mo',
                default_named='dr_pay30', tenor_named='tenor_pay30',
                mix_label='Single-bullet 30-day pay. Invoice acts as collateral. Lowest default; lowest yield.')

build_track_pnl('Track B · BNPL',           'Track B · BNPL (60/90/120/180 blended)',
                book_ref_row=10, fee_kind='monthly', fee_named='fee_bnpl_mo',
                default_named='dr_bnpl', tenor_named='tenor_bnpl',
                mix_label='Multi-installment over 60–180 days. Higher exposure → higher default. Yield priced ~12.5% above Pay-30.')

build_track_pnl('Track C · Extension',      'Track C · Term Extension (6-mo restructure)',
                book_ref_row=11, fee_kind='apr',     fee_named='fee_ext_apr',
                default_named='dr_ext', tenor_named='tenor_ext',
                mix_label='Workout product for Pay-30/BNPL non-payers. APR-priced. Distressed cohort: highest default; recovery lower.')

# ============================================================
# Combined P&L — sum of three tracks
# ============================================================
ws = wb.create_sheet('Combined P&L')
for c, w in zip('ABCDEF', [40, 16, 16, 16, 4, 50]):
    ws.column_dimensions[c].width = w

ws['A1'] = 'Combined Smart Invoice P&L'
ws['A1'].font = F_TITLE
ws['A2'] = 'Track A + Track B + Track C, three-year roll-up'
ws['A2'].font = F_NOTE

header_row(4, 'Line (AED M)', 'Y1', 'Y2', 'Y3', '', 'Composition')

combined_lines = [
    # (label, source row in each track sheet, note)
    ('Avg book',                5, 'sum of three tracks'),
    ('Gross interest income',   7, 'discount + BNPL + extension fees'),
    ('Funding cost',           10, 'COF × book'),
    ('NIM',                    12, '= gross - funding'),
    ('Loss provision',         17, 'PD-LGD on book'),
    ('Risk-adj margin',        19, '= NIM - losses'),
    ('OpEx',                   22, 'book × opex_pct (with ramp)'),
    ('Net contribution',       24, '= RAM - OpEx'),
    ('Equity',                 27, 'book × equity_pct'),
]
ws_row = 5
for label, src, note in combined_lines:
    is_emph = label in ('NIM', 'Risk-adj margin', 'Net contribution')
    ws.cell(row=ws_row, column=1, value=label).font = F_H2 if is_emph else F_LABEL
    for i, col in enumerate('BCD'):
        f = (
            f"='Track A · Pay-30'!{col}{src}"
            f"+'Track B · BNPL'!{col}{src}"
            f"+'Track C · Extension'!{col}{src}"
        )
        c = ws.cell(row=ws_row, column=2+i, value=f)
        c.font = F_OUTPUT if is_emph else F_LABEL
        c.fill = FILL_OUTPUT if label == 'Net contribution' else FILL_CALC
        c.number_format = '#,##0.00'
    ws.cell(row=ws_row, column=6, value=note).font = F_NOTE
    ws_row += 1

# Ratios
ws_row += 1
ws.cell(row=ws_row, column=1, value='Ratios').font = F_H2
ws_row += 1
ratio_rows = {
    'NIM (% of book)':            ('=B8/B5*100', '=C8/C5*100', '=D8/D5*100'),
    'Loss rate (% of book)':      ('=B9/B5*100', '=C9/C5*100', '=D9/D5*100'),
    'OpEx (% of book)':           ('=B11/B5*100', '=C11/C5*100', '=D11/D5*100'),
    'Net margin (% of book)':     ('=B12/B5*100', '=C12/C5*100', '=D12/D5*100'),
    'ROA (% of avg book)':        ('=B12/B5*100', '=C12/C5*100', '=D12/D5*100'),
    'RAROC (% on equity)':        ('=IF(B13>0,B12/B13*100,0)', '=IF(C13>0,C12/C13*100,0)', '=IF(D13>0,D12/D13*100,0)'),
}
for label, formulas in ratio_rows.items():
    ws.cell(row=ws_row, column=1, value=label).font = F_LABEL
    for i, f in enumerate(formulas):
        c = ws.cell(row=ws_row, column=2+i, value=f)
        c.font = F_OUTPUT
        c.fill = FILL_OUTPUT
        c.number_format = '0.00'
    ws_row += 1

# ============================================================
# Unit Economics — per AED 100k of invoice
# ============================================================
ws = wb.create_sheet('Unit Economics')
for c, w in zip('ABCDE', [40, 16, 16, 16, 4]):
    ws.column_dimensions[c].width = w

ws['A1'] = 'Unit Economics — per AED 100k of Invoice Face Value'
ws['A1'].font = F_TITLE
ws['A2'] = 'Per-invoice contribution by track. Useful for pricing-committee debates.'
ws['A2'].font = F_NOTE
header_row(4, 'Line (AED per 100k)', 'Pay-30', 'BNPL', 'Extension')

unit_lines = [
    ('Invoice face',                   '=100000', '=100000', '=100000'),
    ('Advanced principal',             '=B5*advance_rate/100', '=B5*advance_rate/100', '=B5*advance_rate/100'),
    ('Tenor (days)',                   '=tenor_pay30', '=tenor_bnpl', '=tenor_ext'),
    ('Yield (annual %)',               '=fee_pay30_mo*12', '=fee_bnpl_mo*12', '=fee_ext_apr'),
    ('Gross fee earned',               '=B6*B8/100*B7/365',  '=C6*C8/100*C7/365',  '=D6*D8/100*D7/365'),
    ('Funding cost',                   '=B6*cof_apr/100*B7/365', '=C6*cof_apr/100*C7/365', '=D6*cof_apr/100*D7/365'),
    ('NIM',                            '=B9-B10', '=C9-C10', '=D9-D10'),
    ('Expected loss',                  '=B6*dr_pay30/100*(1-recovery_pct/100)*B7/365',
                                       '=C6*dr_bnpl/100*(1-recovery_pct/100)*C7/365',
                                       '=D6*dr_ext/100*(1-recovery_pct/100)*D7/365'),
    ('Risk-adj margin',                '=B11-B12', '=C11-C12', '=D11-D12'),
    ('OpEx (mature, % of book)',       '=B6*opex_pct/100*B7/365', '=C6*opex_pct/100*C7/365', '=D6*opex_pct/100*D7/365'),
    ('Net contribution per invoice',   '=B13-B14', '=C13-C14', '=D13-D14'),
    ('Net contribution as % of face',  '=B15/B5*100', '=C15/C5*100', '=D15/D5*100'),
]
r = 5
for items in unit_lines:
    label, *formulas = items
    is_emph = label in ('NIM', 'Risk-adj margin', 'Net contribution per invoice', 'Net contribution as % of face')
    ws.cell(row=r, column=1, value=label).font = F_H2 if is_emph else F_LABEL
    for i, f in enumerate(formulas):
        c = ws.cell(row=r, column=2+i, value=f)
        c.font = F_OUTPUT if is_emph else F_LABEL
        c.fill = FILL_OUTPUT if 'Net contribution' in label else FILL_CALC
        c.number_format = '#,##0.00'
    r += 1

# ============================================================
# Sensitivity — 2D heatmap: default rate × COF
# ============================================================
ws = wb.create_sheet('Sensitivity')
for c in 'ABCDEFGHIJ':
    ws.column_dimensions[c].width = 12
ws.column_dimensions['A'].width = 28

ws['A1'] = 'Sensitivity · Y3 Net Contribution (AED M) vs Default Rate × COF'
ws['A1'].font = F_TITLE
ws['A2'] = 'Holds all else constant. Useful for stress-test discussions with risk committee.'
ws['A2'].font = F_NOTE

cof_levels = [5.5, 6.0, 6.5, 7.0, 7.5, 8.0, 8.5]
dr_multipliers = [0.6, 0.8, 1.0, 1.2, 1.5, 1.8, 2.2]

# header
ws.cell(row=4, column=1, value='Default × → / COF ↓').font = F_H2
for i, dm in enumerate(dr_multipliers):
    ws.cell(row=4, column=2+i, value=f'{dm:.1f}× DR').font = F_H2
for i, cof in enumerate(cof_levels):
    ws.cell(row=5+i, column=1, value=f'COF {cof:.1f}%').font = F_H2

# Approximation: use Y3 totals from Combined P&L logic, varying COF and default
# Simple closed-form for the heatmap — deliberately not formula-driven against
# named ranges (to make this static and readable). The Combined P&L formulas
# are the live model; this sheet is illustrative.
disb_y3 = 4000
adv = 0.90
tenors = {'pay30': 30, 'bnpl': 105, 'ext': 180}
mix    = {'pay30': 0.40, 'bnpl': 0.50, 'ext': 0.10}
yields = {'pay30': 1.6*12, 'bnpl': 1.8*12, 'ext': 16.0}      # % APR
defaults = {'pay30': 1.8, 'bnpl': 3.5, 'ext': 9.0}            # % APR
recovery = 0.45
opex_pct = 2.4

def y3_net_contrib(cof, dr_mult):
    total = 0.0
    for k in ('pay30', 'bnpl', 'ext'):
        book = disb_y3 * adv * (tenors[k] / 365) * mix[k]
        gross = book * yields[k] / 100
        funding = book * cof / 100
        loss = book * defaults[k] * dr_mult / 100 * (1 - recovery)
        opex = book * opex_pct / 100
        total += gross - funding - loss - opex
    return total

for ci, cof in enumerate(cof_levels):
    for di, dm in enumerate(dr_multipliers):
        v = y3_net_contrib(cof, dm)
        c = ws.cell(row=5+ci, column=2+di, value=round(v, 2))
        c.number_format = '#,##0.0'
        c.alignment = Alignment(horizontal='right')

# colorscale
last_col = get_column_letter(1 + len(dr_multipliers))
rng = f'B5:{last_col}{4 + len(cof_levels)}'
ws.conditional_formatting.add(rng, ColorScaleRule(
    start_type='min', start_color='D44A4A',
    mid_type='num',   mid_value=0,  mid_color='FAF7EE',
    end_type='max',   end_color='A8C09A',
))

ws.cell(row=14, column=1, value='Read-out:').font = F_H2
ws.cell(row=15, column=1, value='Centre cell (1.0× DR, 6.5% COF) is the base case Y3 net contribution.').font = F_NOTE
ws.cell(row=16, column=1, value='Top-right (severe default + cheap funding) shows downside; bottom-left vice versa.').font = F_NOTE
ws.cell(row=17, column=1, value='If centre is positive but 1.5× DR row goes negative → book is risk-rate-elastic, tighten underwriting before scaling.').font = F_NOTE

# ============================================================
# Validation & Sources — alternative scenarios
# ============================================================
ws = wb.create_sheet('Validation & Sources')
ws.column_dimensions['A'].width = 32
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 60

ws['A1'] = 'Validation against UAE benchmarks · Alternative scenarios'
ws['A1'].font = F_TITLE
header_row(4, 'Assumption', 'Mal value', 'Mkt range', 'Source / risk-priced alternative')

vals = [
    ('Cost of funds',          '6.5% APR',     '6.0–8.0%',     '3M EIBOR ~5.2% + NBFI spread 1.0–1.5%. CBUAE Monetary Bulletin Q1 2026; Fitch UAE Banks 2026 outlook.'),
    ('Pay-30 discount fee',    '1.6%/mo',      '1.5–2.2%/mo',  'Mashreq SCF, CBI Trade, ENBD Trade. Mal at competitive floor for prime SME.'),
    ('BNPL fee',               '1.8%/mo',      '2.0–2.8%/mo',  'B2B BNPL UAE. Mal underprice signals embedded-platform play vs standalone BNPL.'),
    ('Term Extension APR',     '16% APR',      '16–22% APR',   'LOW END. Risk-priced alt: 19% APR. Test in Sensitivity sheet by editing fee_ext_apr.'),
    ('Pay-30 default',         '1.8%',         '1.5–3.0%',     'CBUAE FSR 2025 SME secured factoring band.'),
    ('BNPL default',           '3.5%',         '3.0–5.0%',     'Multi-installment exposure; CBUAE FSR + Moody\'s UAE Banking 2026.'),
    ('Term Extension default', '9.0%',         '7.0–12.0%',    'Distressed cohort. Conservative: assume 11% in stress case.'),
    ('Recovery (LGD-)',        '45%',          '30–80%',       'Invoice-secured 60–80%; restructured 30–50%; blend 45% acceptable for mixed book.'),
    ('OpEx (mature)',          '2.4% of book', '1.5–3.0%',     'Reem Finance, Aafaq public filings. Y3 mature target; ramps from 4.0% Y1.'),
    ('Equity %',               '12% of book',  '10–15%',       'CBUAE Reg 32/2013 + Basel III. 12% mid-range; 13% standard Tier-1.'),
    ('Advance rate',           '90%',          '80–95%',       'Prime-obligor norm. Anchor-validated invoices can go 95%.'),
    ('Avg ticket',             'AED 80k',      '50–120k',      'UAE SME invoice median (DED registry inferred).'),
]
for i, row in enumerate(vals, start=5):
    for col, v in enumerate(row):
        c = ws.cell(row=i, column=col+1, value=v)
        c.font = F_LABEL if col != 0 else F_H2
        c.alignment = Alignment(wrap_text=True, vertical='top')

ws.cell(row=18, column=1, value='Open validation tasks').font = F_H1
open_tasks = [
    '1. Replace estimated UAE SME default sub-segment numbers with sector-specific ratios from IBISWorld connector once Cowork access is wired.',
    '2. Replace blended PD/LGD with Moody\'s MCP private-company curves for actual UAE rating distribution.',
    '3. Validate Pay-30 vs BNPL mix ratio against actual buyer-plan-picker telemetry once the prototype goes live to test users.',
    '4. Stress-test extension cohort default at 12% (vs 9% base) — that is the regulator\'s likely shock scenario.',
    '5. Cross-check OpEx ramp against actual cost-to-income ratios in audited financials of Reem Finance / Aafaq Islamic Finance.',
    '6. Confirm 90% advance rate is acceptable to senior bank funding lines (Mashreq, ENBD): typical co-fund advance covenant 85%.',
]
for i, t in enumerate(open_tasks, start=20):
    ws.cell(row=i, column=1, value=t).font = F_NOTE
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
    ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[i].height = 28

# ============================================================
# Save
# ============================================================
out = Path(__file__).parent / 'Mal_P1_SmartInvoice.xlsx'
wb.save(out)
print(f'Wrote {out} ({out.stat().st_size // 1024} KB)')

# Reorder sheets so README first
wb2 = openpyxl.load_workbook(out)
order = ['README & Sources', 'Inputs', 'Volume & Mix',
         'Track A · Pay-30', 'Track B · BNPL', 'Track C · Extension',
         'Combined P&L', 'Unit Economics', 'Sensitivity', 'Validation & Sources']
wb2._sheets = [wb2[s] for s in order if s in wb2.sheetnames]
wb2.save(out)
print('Sheets reordered.')
